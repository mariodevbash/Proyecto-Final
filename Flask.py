from flask import Flask, render_template, request, flash
import xlwt
from xlwt import Workbook
import pandas as pd
import datetime
import xlsxwriter
import os
import fpdf
from fpdf import FPDF
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from pymongo import MongoClient
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)

client = MongoClient("mongodb://127.0.0.1:27017")
db = client.Comandos
todos = db.ComandosEnviados

usuario = ""
correo = ""
comando = ""

@app.errorhandler(404)
def not_found(error):
    return render_template("Error.html")

@app.route("/", methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
        print("Ciente conectado con exito")
    elif request.form.get('Ingresar') == 'Ingresar':
        global usuario
        global correo
        usuario = request.form['Usuario']
        correo = request.form['Correo']
        print("Se ha detectado una nueva conexion.\nUsuario: "+usuario+"\nCorreo: "+correo)
        return render_template("index.html")
    return render_template("User.html")

@app.route("/Menu", methods=['GET', 'POST'])
def Menu():
    return render_template("index.html")

@app.route("/Comandos", methods=['GET', 'POST'])
def Comandos():
    if request.form.get('Enviar') == 'Enviar':
        global comando
        comando = request.form['comando']
        now = datetime.datetime.now()
        newDirName = now.strftime("%H:%M:%S %d-%m-%Y")
        database_entry={'Usuario:':usuario, 'Comado:':comando, 'Hora y Fecha:':newDirName, 'Ip Del Servidor:':request.remote_addr}
        todos.insert_one(database_entry)
        os.system(comando+" > Respuesta.txt")
        with open ("Respuesta.txt", "r") as myfile:
            data=myfile.read().replace('\n', ' ')
            myfile.close()
        if not data:
            data="Comando invalido, intenta con uno nuevo!"
        else: data=data
        return render_template("Comandos.html", data=data)
    elif request.form.get('Excel') == 'Excel':
        datos = pd.read_csv('Respuesta.txt',error_bad_lines=False,engine="python",index_col=False,header=None)
        datos.to_excel("Excel.xlsx", index=False, header=False)
        Fuente = request.form.get("Fuente")
        Tamaño = request.form.get("Tamaño")
        Color = request.form.get("Color")
        if Color == "Negro":
            Colorfont = "000000"
        elif Color == "Rojo":
            Colorfont = "DB0000"
        elif Color == "Verde":
            Colorfont = "0B6623"
        file = 'Excel.xlsx'
        wb = load_workbook(filename=file)
        ws = wb['Sheet1']
        colorfuente = Font(name=Fuente, color=Colorfont, size=Tamaño)
        for cell in ws["A"]:
            cell.font = colorfuente
        wb.save(filename=file)
        print("\nEl archivo excel se ha creado con exito!")
        # Iniciamos los parámetros del script
        remitente = 'solariumdelvallebr@gmail.com'
        destinatarios = correo
        asunto = '[FLASK]Resultado del comando '+comando+' en formato Excel'
        cuerpo = 'EL siguiente archivo contiene el resultado del comando solicitado!'
        ruta_adjunto = 'Excel.xlsx'
        nombre_adjunto = 'Excel.xlsx'
		# Creamos el objeto mensaje
        mensaje = MIMEMultipart()
 
				# Establecemos los atributos del mensaje
        mensaje['From'] = remitente
        mensaje['To'] = destinatarios
        mensaje['Subject'] = asunto
 
		# Agregamos el cuerpo del mensaje como objeto MIME de tipo texto
        mensaje.attach(MIMEText(cuerpo, 'plain'))
 
		# Abrimos el archivo que vamos a adjuntar
        archivo_adjunto = open(ruta_adjunto, 'rb')
 
		# Creamos un objeto MIME base
        adjunto_MIME = MIMEBase('application', 'octet-stream')
		# Y le cargamos el archivo adjunto
        adjunto_MIME.set_payload((archivo_adjunto).read())
		# Codificamos el objeto en BASE64
        encoders.encode_base64(adjunto_MIME)
	    # Agregamos una cabecera al objeto
        adjunto_MIME.add_header('Content-Disposition', "attachment; filename= %s" % nombre_adjunto)
		# Y finalmente lo agregamos al mensaje
        mensaje.attach(adjunto_MIME)
		 
		# Creamos la conexión con el servidor
        sesion_smtp = smtplib.SMTP('smtp.gmail.com', 587)
		 
		# Ciframos la conexión
        sesion_smtp.starttls()
		
		# Iniciamos sesión en el servidor
        sesion_smtp.login('solariumdelvallebr@gmail.com','Solariumdelvalle')
		
		# Convertimos el objeto mensaje a texto
        texto = mensaje.as_string()

		# Enviamos el mensaje
        sesion_smtp.sendmail(remitente, destinatarios, texto)

		# Cerramos la conexión
        sesion_smtp.quit()
        return render_template("ExcelE.html", datos=datos)
    elif request.form.get('Pdf') == 'Pdf':
        with open ("Respuesta.txt", "r", encoding="utf-8") as myfile:
            data=myfile.read().replace('\n', ' ')
            myfile.close()
        Fuente = request.form.get("Fuente")
        Tamaño = request.form.get("Tamaño")
        Orientacion = request.form.get("Orientacion")
        Color = request.form.get("Color")
        if Color == "Negro":
            R = 0
            G = 0
            B = 0
        elif Color == "Rojo":
            R = 255
            G = 0
            B = 0
        elif Color == "Verde":
            R = 0
            G = 255
            B = 0
        pdf=FPDF()
        pdf.add_page()
        pdf.set_font(Fuente,"",int(Tamaño))
        pdf.set_text_color(R,G,B)
        pdf.set_margins(10,10,10)
        pdf.multi_cell(0, 6, data, 0, Orientacion)
        pdf.ln(h="")
        pdf.output("PDF.pdf","f")
        print("\nEl archivo pdf se ha creado con exito!")
        # Iniciamos los parámetros del script
        remitente = 'solariumdelvallebr@gmail.com'
        destinatarios = correo
        asunto = '[FLASK]Resultado del comando '+comando+' en formato PDF'
        cuerpo = 'EL siguiente archivo contiene el resultado del comando solicitado!'
        ruta_adjunto = 'PDF.pdf'
        nombre_adjunto = 'PDF.pdf'
		# Creamos el objeto mensaje
        mensaje = MIMEMultipart()
 
		# Establecemos los atributos del mensaje
        mensaje['From'] = remitente
        mensaje['To'] = destinatarios
        mensaje['Subject'] = asunto
 
		# Agregamos el cuerpo del mensaje como objeto MIME de tipo texto
        mensaje.attach(MIMEText(cuerpo, 'plain'))
 
		# Abrimos el archivo que vamos a adjuntar
        archivo_adjunto = open(ruta_adjunto, 'rb')
 
		# Creamos un objeto MIME base
        adjunto_MIME = MIMEBase('application', 'octet-stream')
		# Y le cargamos el archivo adjunto
        adjunto_MIME.set_payload((archivo_adjunto).read())
		# Codificamos el objeto en BASE64
        encoders.encode_base64(adjunto_MIME)
	    # Agregamos una cabecera al objeto
        adjunto_MIME.add_header('Content-Disposition', "attachment; filename= %s" % nombre_adjunto)
		# Y finalmente lo agregamos al mensaje
        mensaje.attach(adjunto_MIME)
		 
		# Creamos la conexión con el servidor
        sesion_smtp = smtplib.SMTP('smtp.gmail.com', 587)
		 
		# Ciframos la conexión
        sesion_smtp.starttls()
		
		# Iniciamos sesión en el servidor
        sesion_smtp.login('solariumdelvallebr@gmail.com','Solariumdelvalle')
		
		# Convertimos el objeto mensaje a texto
        texto = mensaje.as_string()

		# Enviamos el mensaje
        sesion_smtp.sendmail(remitente, destinatarios, texto)

		# Cerramos la conexión
        sesion_smtp.quit()
        return render_template("PdfE.html", data=data)
    return render_template("Comandos.html")

@app.route("/Registro", methods=['GET', 'POST'])
def Registro():
    registro = open("Registros.txt", "w")
    for doc in todos.find():
        registro.writelines(str(doc)+"\n")
    registro.close()
    with open ("Registros.txt", "r", encoding="utf-8") as myfile:
        data=myfile.readlines()
        myfile.close()
    return render_template("Registro.html", data=data)

if __name__ == "__main__":
    app.run(debug=True,host='0.0.0.0')